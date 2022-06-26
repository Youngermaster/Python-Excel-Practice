import openpyxl as xl
import os
from copy import copy

ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
DOCUMENT_PATH = os.path.join(ROOT_DIR, 'documents')


def get_worksheets():
    print("get_worksheets")
    # opening the source excel file
    filename = os.path.join(DOCUMENT_PATH, 'trading.xlsx')
    wb1 = xl.load_workbook(filename)
    print("Sheet names:", wb1.sheetnames)


def get_data_by_columns():
    print("get_data_by_columns")
    filename = os.path.join(DOCUMENT_PATH, 'trading.xlsx')
    workbook = xl.load_workbook(filename)
    worksheet = workbook['TD']

    # Create a dictionary of column names
    ColNames = {}
    Current = 0
    for COL in worksheet.iter_cols(1, worksheet.max_column):
        ColNames[COL[0].value] = Current
        Current += 1

    # Now you can access by column name
    # (My data has a column named 'Dogs')
    # If we want rang
    # for row_cells in worksheet.iter_rows(min_row=1, max_row=4):
    for row_cells in worksheet.iter_rows():
        print(row_cells[ColNames['Row 4']].value)


def filter_by_columns():
    print("filter_by_columns")
    filename = os.path.join(DOCUMENT_PATH, 'trading.xlsx')
    workbook = xl.load_workbook(filename)
    worksheet = workbook['TD']

    # Create a dictionary of column names
    ColNames = {}
    Current = 0
    for COL in worksheet.iter_cols(1, worksheet.max_column):
        ColNames[COL[0].value] = Current
        Current += 1

    # Now you can access by column name
    # (My data has a column named 'Dogs')
    for row_cells in worksheet.iter_rows():
        if row_cells[ColNames['Filter']].value == True:
            print(row_cells[ColNames['Row 4']].value)


def filter_by_row():
    print("filter_by_row")


def get_data_in_range():
    print("get_data_in_range")


def write_data_in_range():
    print("write_data_in_range")


def copy_worbook():
    # opening the source excel file
    filename = os.path.join(DOCUMENT_PATH, 'trading.xlsx')
    wb1 = xl.load_workbook(filename)
    ws1 = wb1.worksheets[0]

    # opening the destination excel file
    filename1 = os.path.join(DOCUMENT_PATH, 'test.xlsx')
    wb2 = xl.load_workbook(filename1)
    ws2 = wb2.active

    # copying the cell values from source
    # excel file to destination excel file
    for row in ws1.rows:
        for cell in row:
            new_cell = ws2.cell(row=cell.row, column=cell.col_idx,
                                value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # saving the destination excel file
    wb2.save(str(filename1))


if __name__ == "__main__":
    get_worksheets()
    get_data_by_columns()
    filter_by_columns()
